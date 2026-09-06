# 繁工AI 本地解析工作台 - 竣工资料自动组卷（v0.1.24）
# 依据 v3.7 口径：资料生成按项目进度推进，生成即存档；组卷按竣工归档目录归卷，
# 卷内目录导出 Excel，缺项列出待补，最终整卷打包 zip 供打印签字。
import datetime
import io
import json
import os
import zipfile

from . import config
from . import docgen

# 卷宗结构：竣工归档常用类目（可扩展）
VOLUMES = [
    {"no": "01", "name": "开工与施工方案", "docs": ["施工方案"]},
    {"no": "02", "name": "吊装与安全专项方案", "docs": ["吊装方案"]},
    {"no": "03", "name": "技术交底", "docs": ["技术交底"]},
    {"no": "04", "name": "设备开箱与隐蔽验收", "docs": ["开箱验收记录", "隐蔽工程验收记录"]},
    {"no": "05", "name": "施工进度计划", "docs": ["施工计划"]},
    {"no": "06", "name": "施工记录", "docs": ["施工日志"]},
    {"no": "07", "name": "变更与签证", "docs": ["设计变更", "货损报告"]},
    {"no": "08", "name": "竣工验收", "docs": ["竣工资料"]},
]


def _gen_dir() -> str:
    d = os.path.join(config.DATA_DIR, "generated_docs")
    os.makedirs(d, exist_ok=True)
    return d


def _save_generated(doc_type: str, content: bytes, devices: list = None, workshops: list = None) -> str:
    """生成文档落盘存档（v0.1.24）：data/generated_docs/{类型}/繁工AI_{类型}_{时间}.docx
    v0.1.39：保存后自动登记设备/车间关联。"""
    tdir = os.path.join(_gen_dir(), doc_type)
    os.makedirs(tdir, exist_ok=True)
    fname = f"繁工AI_{doc_type}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.docx"
    path = os.path.join(tdir, fname)
    with open(path, "wb") as fh:
        fh.write(content)
    # v0.1.39：自动登记资料关联
    try:
        from . import doc_relations as _dr
        _dr.register_doc(f"{doc_type}/{fname}", doc_type, file_path=path,
                         devices=devices, workshops=workshops, source="generated")
    except Exception:  # noqa: BLE001
        pass
    return path


def _list_generated() -> dict:
    """doc_type -> [文件路径]（按 mtime 倒序）"""
    out = {}
    root = _gen_dir()
    if not os.path.isdir(root):
        return out
    for t in os.listdir(root):
        tdir = os.path.join(root, t)
        if not os.path.isdir(tdir):
            continue
        files = []
        for f in os.listdir(tdir):
            if f.endswith(".docx"):
                p = os.path.join(tdir, f)
                files.append((os.path.getmtime(p), p))
        files.sort(reverse=True)
        out[t] = [p for _, p in files]
    return out


def archive_status() -> dict:
    """每卷：已有资料、缺失类型、卷内清单；统计齐全度。"""
    gen = _list_generated()
    volumes = []
    total_have, total_need = 0, 0
    for v in VOLUMES:
        items = []
        missing = []
        for t in v["docs"]:
            files = gen.get(t, [])
            if files:
                items.append({"doc_type": t, "count": len(files),
                              "latest": os.path.basename(files[0]),
                              "ts": datetime.datetime.fromtimestamp(os.path.getmtime(files[0])).strftime("%Y-%m-%d %H:%M")})
                total_have += 1
            else:
                missing.append(t)
                total_need += 1
        volumes.append({"no": v["no"], "name": v["name"],
                        "docs": v["docs"], "items": items, "missing": missing,
                        "ready": not missing})
    return {
        "volumes": volumes,
        "generated_total": sum(len(f) for f in gen.values()),
        "ready_volumes": sum(1 for v in volumes if v["ready"]),
        "total_volumes": len(VOLUMES),
        "have_types": total_have,
        "need_types": total_have + total_need,
        "latest_build": datetime.datetime.now().isoformat(),
    }


def export_archive() -> bytes:
    """按卷组织生成 zip：卷目录/文件 + 卷内目录.xlsx。"""
    gen = _list_generated()
    buf = io.BytesIO()
    # 卷内目录
    rows = [["卷号", "卷名", "资料类型", "文件", "生成时间", "状态"]]
    for v in VOLUMES:
        for t in v["docs"]:
            files = gen.get(t, [])
            if files:
                for p in files:
                    ts = datetime.datetime.fromtimestamp(os.path.getmtime(p)).strftime("%Y-%m-%d %H:%M")
                    rows.append([v["no"], v["name"], t, os.path.basename(p), ts, "已归档"])
            else:
                rows.append([v["no"], v["name"], t, "—", "—", "缺失（待生成）"])
    xlsx = _make_xlsx(rows)
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("00_卷内目录.xlsx", xlsx)
        zf.writestr("README.txt",
                    "繁工AI 竣工资料归档包\n"
                    "导出时间：%s\n"
                    "说明：按卷宗结构组织；缺失项见卷内目录『缺失（待生成）』行，"
                    "在⑧资料生成计划中补齐后重新导出。签字由人工下载打印完成。\n"
                    % datetime.datetime.now().isoformat())
        for v in VOLUMES:
            for t in v["docs"]:
                for p in gen.get(t, []):
                    arc = f"{v['no']}_{v['name']}/{os.path.basename(p)}"
                    zf.write(p, arc)
    return buf.getvalue()


def _make_xlsx(rows: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "卷内目录"
    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = Font(name="微软雅黑", size=10, bold=(i == 1))
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col, w in zip("ABCDEF", [10, 22, 20, 40, 18, 16]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ==================== v0.1.48 竣工资料自动组卷增强 ====================

# 专业分类映射（资料类型 → 专业）
PROFESSION_MAP = {
    "施工方案": "工艺设备",
    "吊装方案": "工艺设备",
    "技术交底": "综合",
    "开箱验收记录": "工艺设备",
    "隐蔽工程验收记录": "工艺设备",
    "施工计划": "综合",
    "施工日志": "综合",
    "设计变更": "综合",
    "货损报告": "综合",
    "竣工资料": "综合",
    "安全交底": "安全",
    "质量验收记录": "质量",
    "压力试验记录": "工艺管道",
    "管道焊接记录": "工艺管道",
    "电气调试记录": "电气",
    "仪表调试记录": "仪表",
    "设备试运转记录": "工艺设备",
}

# 组卷规则：卷 → 专业 → 车间 → 设备
def _get_profession(doc_type: str) -> str:
    """获取资料所属专业。"""
    return PROFESSION_MAP.get(doc_type, "其他")


def _get_doc_workshop(file_path: str) -> str:
    """从资料关联中获取车间（v0.1.39 doc_relations）。"""
    try:
        from . import doc_relations as _dr
        rels = _dr.list_relations()
        fname = os.path.basename(file_path)
        for r in rels:
            if r.get("file_name") == fname or r.get("doc_id", "").endswith(fname):
                ws = r.get("workshops") or []
                if ws:
                    return ws[0]
    except Exception:  # noqa: BLE001
        pass
    return "未分类"


def _get_doc_devices(file_path: str) -> list:
    """从资料关联中获取设备列表。"""
    try:
        from . import doc_relations as _dr
        rels = _dr.list_relations()
        fname = os.path.basename(file_path)
        for r in rels:
            if r.get("file_name") == fname or r.get("doc_id", "").endswith(fname):
                return r.get("devices") or []
    except Exception:  # noqa: BLE001
        pass
    return []


def archive_status_enhanced() -> dict:
    """v0.1.48：增强版归档状态（多级分类：卷→专业→车间→设备）。"""
    gen = _list_generated()
    volumes = []
    total_files = 0

    for v in VOLUMES:
        volume_data = {
            "no": v["no"], "name": v["name"],
            "professions": {},  # 专业 → {车间 → {设备 → [文件]}}
            "items": [], "missing": [],
            "file_count": 0,
        }
        for t in v["docs"]:
            files = gen.get(t, [])
            if files:
                profession = _get_profession(t)
                if profession not in volume_data["professions"]:
                    volume_data["professions"][profession] = {}
                for p in files:
                    workshop = _get_doc_workshop(p)
                    devices = _get_doc_devices(p)
                    if workshop not in volume_data["professions"][profession]:
                        volume_data["professions"][profession][workshop] = {}
                    dev_key = ",".join(devices) if devices else "通用"
                    if dev_key not in volume_data["professions"][profession][workshop]:
                        volume_data["professions"][profession][workshop][dev_key] = []
                    volume_data["professions"][profession][workshop][dev_key].append({
                        "file": os.path.basename(p),
                        "path": p,
                        "doc_type": t,
                        "ts": datetime.datetime.fromtimestamp(os.path.getmtime(p)).strftime("%Y-%m-%d %H:%M"),
                        "devices": devices,
                        "workshop": workshop,
                    })
                    volume_data["file_count"] += 1
                    total_files += 1
                volume_data["items"].append({"doc_type": t, "count": len(files)})
            else:
                volume_data["missing"].append(t)
        volume_data["ready"] = not volume_data["missing"]
        volumes.append(volume_data)

    return {
        "volumes": volumes,
        "total_files": total_files,
        "ready_volumes": sum(1 for v in volumes if v["ready"]),
        "total_volumes": len(VOLUMES),
        "completeness": round(total_files / max(1, sum(len(v["docs"]) for v in VOLUMES)) * 100, 1),
    }


def generate_volume_catalog(volume_no: str = None) -> list:
    """v0.1.48：生成卷内详细目录（含多级分类）。"""
    status = archive_status_enhanced()
    rows = [["卷号", "卷名", "专业", "车间", "设备", "资料类型", "文件名", "生成时间", "状态"]]
    for v in status["volumes"]:
        if volume_no and v["no"] != volume_no:
            continue
        for profession, workshops in v["professions"].items():
            for workshop, devices in workshops.items():
                for dev_key, files in devices.items():
                    for f in files:
                        rows.append([v["no"], v["name"], profession, workshop,
                                     dev_key, f["doc_type"], f["file"], f["ts"], "已归档"])
        # 缺失项
        for t in v["missing"]:
            rows.append([v["no"], v["name"], _get_profession(t), "—", "—", t, "—", "—", "缺失（待生成）"])
    return rows


def check_archive_completeness() -> dict:
    """v0.1.48：归档完整性检查（与v0.1.36 completeness_check联动）。"""
    status = archive_status_enhanced()
    # 检查每卷的缺失项
    missing_by_volume = {}
    for v in status["volumes"]:
        if v["missing"]:
            missing_by_volume[v["no"]] = {
                "name": v["name"],
                "missing_docs": v["missing"],
                "existing_count": v["file_count"],
            }
    # 检查设备级资料完整性（每台设备应有开箱+安装+验收记录）
    device_completeness = {}
    try:
        from . import relations as _rel
        g = _rel.load_relations()
        for dev in g.get("devices", []):
            tag = dev["tag"]
            has_openbox = False
            has_install = False
            has_acceptance = False
            for v in status["volumes"]:
                for profession, workshops in v["professions"].items():
                    for workshop, devices in workshops.items():
                        for dev_key, files in devices.items():
                            if tag in dev_key:
                                for f in files:
                                    if "开箱" in f["doc_type"]:
                                        has_openbox = True
                                    elif "隐蔽" in f["doc_type"]:
                                        has_install = True
                                    elif "竣工" in f["doc_type"] or "验收" in f["doc_type"]:
                                        has_acceptance = True
            missing = []
            if not has_openbox:
                missing.append("开箱验收记录")
            if not has_install:
                missing.append("隐蔽工程验收记录")
            if not has_acceptance:
                missing.append("竣工验收记录")
            if missing:
                device_completeness[tag] = {"missing": missing, "complete": False}
            else:
                device_completeness[tag] = {"missing": [], "complete": True}
    except Exception:  # noqa: BLE001
        pass

    return {
        "overall_completeness": status["completeness"],
        "ready_volumes": status["ready_volumes"],
        "total_volumes": status["total_volumes"],
        "missing_by_volume": missing_by_volume,
        "device_completeness": device_completeness,
        "devices_with_missing": sum(1 for v in device_completeness.values() if not v["complete"]),
        "total_devices": len(device_completeness),
    }


def export_archive_enhanced() -> bytes:
    """v0.1.48：增强版归档导出（多级文件夹结构：卷/专业/车间/设备/文件）。"""
    status = archive_status_enhanced()
    buf = io.BytesIO()

    # 卷内详细目录
    catalog_rows = generate_volume_catalog()
    xlsx = _make_xlsx_enhanced(catalog_rows)

    # 完整性检查报告
    completeness = check_archive_completeness()

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("00_卷内详细目录.xlsx", xlsx)
        zf.writestr("01_归档完整性检查报告.json",
                     json.dumps(completeness, ensure_ascii=False, indent=2))
        zf.writestr("README.txt",
                    "繁工AI 竣工资料归档包（增强版 v0.1.48）\n"
                    "导出时间：%s\n"
                    "整体齐全度：%.1f%%\n"
                    "已就绪卷：%d/%d\n"
                    "目录结构：卷号_卷名/专业/车间/设备/文件\n"
                    "缺失项见 00_卷内详细目录.xlsx 和 01_归档完整性检查报告.json\n"
                    "签字由人工下载打印完成。\n"
                    % (datetime.datetime.now().isoformat(),
                       completeness["overall_completeness"],
                       completeness["ready_volumes"],
                       completeness["total_volumes"]))

        # 按多级结构组织文件
        for v in status["volumes"]:
            for profession, workshops in v["professions"].items():
                for workshop, devices in workshops.items():
                    for dev_key, files in devices.items():
                        for f in files:
                            # 路径：卷号_卷名/专业/车间/设备/文件名
                            safe_dev = dev_key.replace("/", "_").replace("\\", "_")[:50]
                            arc = "%s_%s/%s/%s/%s/%s" % (
                                v["no"], v["name"], profession, workshop, safe_dev, f["file"])
                            zf.write(f["path"], arc)

    return buf.getvalue()


def _make_xlsx_enhanced(rows: list) -> bytes:
    """增强版Excel生成（9列）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "卷内详细目录"
    header_fill = PatternFill(start_color="1E5AA8", end_color="1E5AA8", fill_type="solid")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    missing_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = header_font if i == 1 else Font(name="微软雅黑", size=10)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if i == 1:
                c.fill = header_fill
            elif len(row) > 8 and row[8] and "缺失" in str(row[8]):
                c.fill = missing_fill

    for col, w in zip("ABCDEFGHI", [8, 20, 12, 12, 20, 18, 35, 18, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:I%d" % len(rows)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
